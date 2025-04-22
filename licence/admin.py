from django.contrib import admin,messages
from django.urls import path, reverse
from django.http import HttpResponse # 导入 HttpResponse
from django.shortcuts import redirect
from django.utils.html import format_html
from urllib.parse import quote # 添加这行
from django.utils import timezone # 导入 timezone
from django.db.models import Sum, Count # 导入 Sum 和 Count
from .models import Licence
from orders.models import Order # 导入 Order 模型
import openpyxl # 导入 openpyxl
from openpyxl.utils import get_column_letter # 用于设置列宽
import json # 导入 json 模块
# Register your models here.
@admin.register(Licence)
class LicenceAdmin(admin.ModelAdmin):
    list_display = ("machine_code","type","remark","daily_orders_summary","status_display","type_button")
    save_as_continue = False
    actions = ['export_3d_orders_excel'] # 添加导出 action
    def daily_orders_summary(self, obj):
        """计算并显示今日订单数量和总额"""
        today_str = timezone.now().strftime('%Y/%m/%d')
        # 假设 Order 模型中关联 Licence 的字段是 machineId
        # 并且 Order 模型中有表示价格的字段 price 和表示销售时间的字段 saleTime (DateTimeField)
        orders_today = Order.objects.filter(
            machineId=obj.machine_code,
            saleTime__startswith=today_str # 筛选今天的订单
        )

        # 计算总数和总额
        summary = orders_today.aggregate(
            count=Count('id'),
            total=Sum('price') # 假设价格字段是 price
        )

        count = summary['count'] or 0
        total = summary['total'] or 0.00 # 如果没有订单，总额为0

        return f"数量: {count}, 总额: {total:.2f}" # 格式化输出
        # 设置新列的表头名称
    daily_orders_summary.short_description = "今日出票 (数量/总额)"

    def type_button(self, obj):
        """在列表里显示按钮：若启用则显示“禁用”，若停用则显示“启用”"""
        if obj.enabled == 1:
            return format_html(
                '<a type="button" class="el-button stop-submit el-button--danger el-button--small" data-name="delete_selected" href="{}"><span style="color:white">禁用</span></a>',
                reverse("admin:licence_licence_toggle_enabled", args=[obj.pk])
            )
        else:
            return format_html(
                '<a type="button" class="el-button el-button--primary el-button--small" data-name="add_item" href="{}"><span style="color:white">启用</span></a>',
                reverse("admin:licence_licence_toggle_enabled", args=[obj.pk])
            )
    type_button.short_description = "操作"

    def status_display(self, obj):
        """给状态添加样式"""
        if obj.status == 1:  # 在线
            return format_html('<span class="el-tag el-tag--light">在线</span>')
        else:  # 离线
            return format_html('<span class="el-tag el-tag--danger el-tag--light">离线</span>')
    status_display.short_description = "机器状态"

    def has_add_permission(self,request):
      return False
    def get_urls(self):
        """定义自定义URL，用于切换enabled值"""
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:pk>/toggle_enabled/",  # 注意这里更改了顺序
                self.admin_site.admin_view(self.toggle_enabled),
                name="licence_licence_toggle_enabled",  # 使用下划线而不是破折号
            ),
        ]
        return custom_urls + urls

    def toggle_enabled(self, request, pk, *args, **kwargs):
        """根据当前enabled的值，切换为0或1"""
        licence = Licence.objects.get(pk=pk)
        licence.enabled = 0 if licence.enabled == 1 else 1
        licence.save()
        return redirect(reverse('admin:licence_licence_changelist')) # 返回到上一层列表页
    
       # --- 修改后的导出 Action ---
    def export_3d_orders_excel(self, request, queryset):
        """
        导出所选 Licence 关联的 '3D单选' 订单，聚合相同组合的倍数，并按倍数降序排列。
        """
        selected_machine_codes = list(queryset.values_list('machine_code', flat=True))

        if not selected_machine_codes:
            self.message_user(request, "请先选择要导出订单的机器。", level=messages.WARNING)
            return

        # 筛选符合条件的订单
        orders_to_process = Order.objects.filter(
            machineId__in=selected_machine_codes,
            playMethod='3D单选' # 确保这里的字符串和你的数据一致
        )

        if not orders_to_process.exists():
            self.message_user(request, "未找到所选机器下玩法为 '3D单选' 的订单。", level=messages.INFO)
            return

        # 用于聚合组合和倍数
        aggregated_combinations = {} # 格式: {"1,1,1": total_multiplier}

        # 遍历订单，解析 JSON 并聚合数据
        for order in orders_to_process:
            try:
                # 假设 order.orders 是存储 JSON 字符串的字段名
                orders_list = json.loads(order.orders)
                if not isinstance(orders_list, list):
                    continue # 跳过无效格式

                for item in orders_list:
                    # 确保列表结构符合预期 [玩法, 数字1, 数字2, 数字3, 倍数]
                    if isinstance(item, list) and len(item) == 5:
                        try:
                            # 组合是第2, 3, 4个元素 (索引 1, 2, 3)
                            combination = f"{item[1]},{item[2]},{item[3]}"
                            # 倍数是第5个元素 (索引 4)
                            multiplier = int(item[4])
                            # 累加倍数
                            aggregated_combinations[combination] = aggregated_combinations.get(combination, 0) + multiplier
                        except (ValueError, TypeError, IndexError):
                            # 跳过列表内格式错误或倍数无法转换的情况
                            continue
            except (json.JSONDecodeError, TypeError):
                # 跳过 JSON 解析失败或字段类型不是字符串的情况
                continue

        if not aggregated_combinations:
            self.message_user(request, "处理订单数据后未找到有效的 '3D单选' 组合。", level=INFO)
            return

        # 将聚合结果转换为列表并按倍数降序排序
        sorted_combinations = sorted(
            aggregated_combinations.items(),
            key=lambda item: item[1], # 按倍数 (字典的值) 排序
            reverse=True # 降序
        )

        # 创建 Excel 工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "3D单选组合统计"

        # 定义表头
        headers = ["来源机器", "组合", "总倍数"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = openpyxl.styles.Font(bold=True)
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20 # 设置合适的列宽

        # 确定第一列的显示内容 (来源机器)
        first_licence = queryset.first() # 获取选中的第一个 Licence 对象
        source_display = first_licence.remark or first_licence.machine_code if first_licence else "未知来源"

        # 写入数据行
        row_num = 2
        for combination, total_multiplier in sorted_combinations:
            ws.cell(row=row_num, column=1, value=source_display)
            ws.cell(row=row_num, column=2, value=combination)
            ws.cell(row=row_num, column=3, value=total_multiplier)
            row_num += 1

        # --- 文件名处理和响应头设置 ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        # 1. 创建原始文件名 (包含非 ASCII 字符)
        original_filename = f"{source_display}_3D单选组合统计.xlsx"

        # 2. 创建 ASCII 安全的文件名 (替换或移除特殊字符)
        ascii_filename = "".join(c if ord(c) < 128 else '_' for c in original_filename).replace(' ', '_').replace('/', '_')
        # 确保文件名不为空
        ascii_filename = ascii_filename or "export.xlsx"

               # 使用 quote 替换 urlquote
        utf8_filename = quote(original_filename)

        # 4. 设置 Content-Disposition 头，同时包含 filename 和 filename*
        response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{utf8_filename}'

        wb.save(response)
        return response

    # 设置 Action 在 Admin 界面的显示名称
    export_3d_orders_excel.short_description = "导出3D单选组合统计"